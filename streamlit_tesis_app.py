import streamlit as st
import pandas as pd
import numpy as np
import io
from sentence_transformers import SentenceTransformer, util
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# 🌎 Título e introducción
titulo = "Clasificador de Títulos de Tesis"
st.set_page_config(page_title=titulo, layout="centered")
st.title(titulo)
st.subheader("Creado por la arquitecta María José Duarte Torres")

st.markdown("""
Esta herramienta permite analizar los títulos o subtítulos de un documento de tesis y clasificarlos automáticamente según su nivel de relevancia. Evalúa si se relacionan con los objetivos, el marco teórico y la metodología, si se repiten en otros capítulos, y si podrían resumirse o eliminarse. Finalmente, se genera una clasificación de A (alta relevancia) a E (baja relevancia) con colores y sugerencias.

**Recomendaciones:**
- El archivo Excel debe tener una columna llamada **"Capítulo o título"**.
- Suba archivos `.txt` para los **objetivos**, **marco teórico** y **metodología**.
""")

# 📚 Subida de archivos
archivo_excel = st.file_uploader("Sube el archivo Excel con los capítulos o títulos", type=[".xlsx"])
objetivos_txt = st.file_uploader("Sube el archivo de objetivos.txt", type="txt")
marco_txt = st.file_uploader("Sube el archivo de marco_teorico.txt", type="txt")
metodologia_txt = st.file_uploader("Sube el archivo de metodologia.txt", type="txt")

if st.button("📊 Generar tabla de clasificación"):
    try:
        if not (archivo_excel and objetivos_txt and marco_txt and metodologia_txt):
            st.error("Por favor, sube todos los archivos requeridos.")
            st.stop()

        # 🔖 Leer Excel y textos
        df = pd.read_excel(archivo_excel)
        if "Capítulo o título" not in df.columns:
            st.error("La columna 'Capítulo o título' no está presente en el archivo Excel.")
            st.stop()

        titulos = df["Capítulo o título"].astype(str).tolist()
        objetivos = objetivos_txt.read().decode("utf-8")
        marco = marco_txt.read().decode("utf-8")
        metodologia = metodologia_txt.read().decode("utf-8")

        # 🧬 Modelo de embeddings
        model = SentenceTransformer('all-MiniLM-L6-v2')
        emb_titulos = model.encode(titulos, convert_to_tensor=True)
        emb_obj = model.encode([objetivos], convert_to_tensor=True)
        emb_marco = model.encode([marco], convert_to_tensor=True)
        emb_metodo = model.encode([metodologia], convert_to_tensor=True)

        # 🔄 Calcular similitud y clasificar
        def similitud(x, emb):
            return float(util.cos_sim(x, emb)[0][0])

        relaciones = []
        categorias = []
        marco_rels = []
        metodo_rels = []
        repetidos = []
        resumen = []

        vistos = {}

        for i, emb in enumerate(emb_titulos):
            s_obj = similitud(emb, emb_obj)
            s_marco = similitud(emb, emb_marco)
            s_metodo = similitud(emb, emb_metodo)

            relaciones.append("Sí" if s_obj > 0.3 else "No")
            marco_rels.append("Sí" if s_marco > 0.3 else "No")
            metodo_rels.append("Sí" if s_metodo > 0.3 else "No")

            tit = titulos[i].lower().strip()
            vistos[tit] = vistos.get(tit, 0) + 1
            repetidos.append("Sí" if vistos[tit] > 1 else "No")

            if len(titulos[i]) < 25 or (s_obj < 0.3 and s_marco < 0.3 and s_metodo < 0.3):
                resumen.append("Sí")
            else:
                resumen.append("No")

            score = s_obj + s_marco + s_metodo
            if score > 1.0:
                categorias.append("A")
            elif score > 0.8:
                categorias.append("B")
            elif score > 0.6:
                categorias.append("C")
            elif score > 0.4:
                categorias.append("D")
            else:
                categorias.append("E")

        df["Categoría final (A-E)"] = categorias
        df["¿Relaciona un objetivo?"] = relaciones
        df["¿Es clave para entender metodología/resultados?"] = metodo_rels
        df["¿Aporta al marco teórico?"] = marco_rels
        df["¿Se repite en otro capítulo?"] = repetidos
        df["¿Puede resumirse/eliminarse?"] = resumen

        # 🔖 Guardar con colores
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Clasificación")
            wb = writer.book
            ws = writer.sheets["Clasificación"]

            for i, row in df.iterrows():
                for col in ["¿Relaciona un objetivo?", "¿Es clave para entender metodología/resultados?",
                            "¿Aporta al marco teórico?", "¿Se repite en otro capítulo?",
                            "¿Puede resumirse/eliminarse?"]:
                    cell = ws.cell(row=i + 2, column=df.columns.get_loc(col) + 1)
                    valor = str(row[col]).strip().lower()
                    if valor == "sí":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif valor == "no":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # 🔹 Colorear la categoría
                categoria = row["Categoría final (A-E)"]
                cell = ws.cell(row=i + 2, column=df.columns.get_loc("Categoría final (A-E)") + 1)
                colores = {
                    "A": "B7E1CD",
                    "B": "D9EAD3",
                    "C": "FFF2CC",
                    "D": "FCE5CD",
                    "E": "F4CCCC"
                }
                if categoria in colores:
                    cell.fill = PatternFill(start_color=colores[categoria], end_color=colores[categoria], fill_type="solid")

        st.success("✅ Clasificación generada con éxito.")
        st.download_button("📥 Descargar archivo clasificado", data=output.getvalue(), file_name="contenido_tesis_clasificado.xlsx")

        st.image("ejemplo_resultado.png", caption="Ejemplo de clasificación generada", use_column_width=True)

    except Exception as e:
        st.error(f"Ocurrió un error inesperado: {e}")
